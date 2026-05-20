from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Expense
from .forms import ExpenseForm

from reportlab.pdfgen import canvas

from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font


@login_required
def expense_list(request):

    expenses = Expense.objects.filter(
        user=request.user
    ).order_by('-date')

    category_query = request.GET.get('category')
    payment_query = request.GET.get('payment')

    if category_query:

        expenses = expenses.filter(
            category__icontains=category_query
        )

    if payment_query:

        expenses = expenses.filter(
            payment_method__icontains=payment_query
        )

    context = {
        'expenses': expenses
    }

    return render(
        request,
        'expenses/expense_list.html',
        context
    )


@login_required
def add_expense(request):

    if request.method == 'POST':

        form = ExpenseForm(request.POST)

        if form.is_valid():

            expense = form.save(commit=False)

            expense.user = request.user

            expense.save()

            return redirect('expense_list')

    else:

        form = ExpenseForm()

    context = {
        'form': form
    }

    return render(
        request,
        'expenses/add_expense.html',
        context
    )


@login_required
def edit_expense(request, pk):

    expense = get_object_or_404(
        Expense,
        pk=pk,
        user=request.user
    )

    if request.method == 'POST':

        form = ExpenseForm(
            request.POST,
            instance=expense
        )

        if form.is_valid():

            form.save()

            return redirect('expense_list')

    else:

        form = ExpenseForm(instance=expense)

    context = {
        'form': form
    }

    return render(
        request,
        'expenses/edit_expense.html',
        context
    )


@login_required
def delete_expense(request, pk):

    expense = get_object_or_404(
        Expense,
        pk=pk,
        user=request.user
    )

    expense.delete()

    return redirect('expense_list')


@login_required
def download_pdf(request):

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename="expenses.pdf"'
    )

    p = canvas.Canvas(response)

    expenses = Expense.objects.filter(
        user=request.user
    )

    y = 800

    p.drawString(100, y, "Expense Report")

    y -= 40

    for expense in expenses:

        line = (
            f"{expense.category} - "
            f"{expense.amount} - "
            f"{expense.date}"
        )

        p.drawString(100, y, line)

        y -= 30

    p.save()

    return response


@login_required
def download_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = 'Expenses'

    headers = [
        'Amount',
        'Category',
        'Payment Method',
        'Description',
        'Date'
    ]

    for col_num, header in enumerate(headers, 1):

        cell = sheet.cell(
            row=1,
            column=col_num
        )

        cell.value = header

        cell.font = Font(bold=True)

    expenses = Expense.objects.filter(
        user=request.user
    )

    row_num = 2

    for expense in expenses:

        sheet.cell(
            row=row_num,
            column=1
        ).value = float(expense.amount)

        sheet.cell(
            row=row_num,
            column=2
        ).value = expense.category

        sheet.cell(
            row=row_num,
            column=3
        ).value = expense.payment_method

        sheet.cell(
            row=row_num,
            column=4
        ).value = expense.description

        sheet.cell(
            row=row_num,
            column=5
        ).value = str(expense.date)

        row_num += 1

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    )

    response['Content-Disposition'] = (
        'attachment; filename=expenses.xlsx'
    )

    workbook.save(response)

    return response